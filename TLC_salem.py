import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from itertools import product
import os
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import json
import sys
import shutil



def get_data_path(filename):
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), filename)
    return os.path.join(os.path.dirname(__file__), filename)


def get_default_json_path():
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, "stations.json")  # Internal default
    return os.path.join(os.path.dirname(__file__), "stations.json")

STATION_FILE = get_data_path("stations.json")

def load_stations():
    global fg, dt
    if not os.path.exists(STATION_FILE):
        try:

            shutil.copy(get_default_json_path(), STATION_FILE)
        except Exception:

            fg = ["Erode", "Jolarpettai"]
            dt = {
                'ED': {'ED','TPMR','PYR','IGR','VZ','UKL','TUP','VNJ','SNO','SUU','IGU','PLMD','CBF','CBE'},
                'JTJ': {'JTJ','TPT','KEY','SLY','DST','DPI','MAP','BDY','BQL','LCR','DSPT','TNT','KPPR','MGSJ','SA','VRPD','DC','MVPM','SGE','ANU','CV','ED'}
            }
            return

    with open(STATION_FILE, "r") as f:
        data = json.load(f)
        fg = data.get("fg", [])
        dt = {k: set(v) for k, v in data.get("dt", {}).items()}

def save_stations():
    with open(STATION_FILE, "w") as f:
        json.dump({
            "fg": fg,
            "dt": {k: list(v) for k, v in dt.items()}
        }, f, indent=4)


pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', None)

fg = ["Erode", "Jolarpettai"]
dt = {
        'ED': {'ED','TPMR','PYR','IGR','VZ','UKL','TUP','VNJ','SNO','SUU','IGU','PLMD','CBF','CBE'},
        'JTJ': {'JTJ','TPT','KEY','SLY','DST','DPI','MAP','BDY','BQL','LCR','DSPT','TNT','KPPR','MGSJ','SA','VRPD','DC','MVPM','SGE','ANU','CV','ED'}
 }

def set_column_widths(excel_path, width=16):
    wb = load_workbook(excel_path)
    ws = wb.active
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
    wb.save(excel_path)

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        file_path_var.set(file_path)

def View_station():
    popup = tk.Toplevel(root)
    popup.title("View/Edit Existing Stations")
    popup.geometry("600x400")


    tk.Label(popup, text="Zone Names (fg list):").pack(pady=5)
    fg_text = tk.Text(popup, height=5, width=70)
    fg_text.pack()
    fg_text.insert("1.0", "\n".join(fg))


    tk.Label(popup, text="Zone Mapping (dt dictionary):").pack(pady=5)
    dt_text = tk.Text(popup, height=10, width=70)
    dt_text.pack()
    dt_str = "\n".join(f"{k}: {', '.join(sorted(v))}" for k, v in dt.items())
    dt_text.insert("1.0", dt_str)

    def save_changes():

        new_fg = fg_text.get("1.0", "end").strip().split("\n")
        fg.clear()
        fg.extend(map(str.strip, new_fg))


        dt_raw = dt_text.get("1.0", "end").strip().split("\n")
        new_dt = {}
        for line in dt_raw:
            if ':' in line:
                key, values = line.split(":", 1)
                key = key.strip()
                values_set = set(map(str.strip, values.split(',')))
                if key in dt:
                    new_dt[key] = values_set
        dt.clear()
        dt.update(new_dt)
        save_stations()

        messagebox.showinfo("Updated", "Changes saved successfully!")
        popup.destroy()

    def cancel():
        popup.destroy()

    tk.Button(popup, text="Save Changes", command=save_changes).pack(pady=5)
    tk.Button(popup, text="Cancel", command=cancel).pack()

def Add_station():
    popup = tk.Toplevel(root)
    popup.title("Add New Station Info")
    popup.geometry("400x250")

    tk.Label(popup, text="Station Zone Name (fg)").pack(pady=5)
    fg_entry = tk.Entry(popup, width=40)
    fg_entry.pack()

    tk.Label(popup, text="Zone Code (dt key)").pack(pady=5)
    key_entry = tk.Entry(popup, width=40)
    key_entry.pack()

    tk.Label(popup, text="Comma-Separated Stations (dt values)").pack(pady=5)
    value_entry = tk.Entry(popup, width=40)
    value_entry.pack()

    def submit():
        zone_name = fg_entry.get().strip()
        dt_key = key_entry.get().strip()
        dt_values_raw = value_entry.get().strip()

        if not zone_name or not dt_key or not dt_values_raw:
            messagebox.showerror("Input Error", "All fields are required.")
            return

        station_set = set(map(str.strip, dt_values_raw.split(',')))

        fg.append(zone_name)
        dt[dt_key] = station_set

        messagebox.showinfo("Success", f"Added zone '{zone_name}' and key '{dt_key}'")
        save_stations()

        popup.destroy()

    def cancel():
        popup.destroy()

    tk.Button(popup, text="Submit", command=submit).pack(pady=10)
    tk.Button(popup, text="Cancel", command=cancel).pack()


def process_file():
    x = file_path_var.get()
    if x == "No file selected":
        messagebox.showerror("Error", "Please select a file.")
        return

    try:
        raw_df = pd.read_excel(x, header=None)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read file:\n{e}")
        return

    n = 0
    for i, r in raw_df.iterrows():
        if "S.No." in r.values or "S.No" in r.values:
            n = int(i)
            break
    else:
        messagebox.showerror("Error", "'S.No.' row not found in file.")
        return

    try:
        df = pd.read_excel(x, skiprows=n)
        df.set_index("S.No.", inplace=True)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process Excel content:\n{e}")
        return




    try:
        for y in range(len(fg)):
            zx = fg[y]
            g = list(dt.values())[y]
            f = list(dt.keys())
            l = list(product(f, g))


            crew_ids_set = set()
            for i, j in l:
                if i != j:
                    op = df[(df["SIGNON STTN"] == i) & (df["SIGNOFF STTN"] == j)]
                    crew_ids_set.update(op["CREW ID"].dropna().unique())

            filtered_df = df[
                (df["CREW ID"].isin(crew_ids_set)) &
                (df["SIGNON STTN"].isin(g)) &
                (df["SIGNOFF STTN"].isin(g))
            ]

            count_sp = (filtered_df["DUTY TYPE"] == "SP").sum()
            count_wr = (filtered_df["DUTY TYPE"] == "WR").sum()

            from_path = os.path.join(os.path.dirname(x), f"From{zx}.xlsx")
            filtered_df.to_excel(from_path)

            wb = load_workbook(from_path)
            ws = wb.active
            next_row = ws.max_row + 2
            ws.cell(row=next_row, column=1, value=f"SP COUNT: {count_sp}")
            ws.cell(row=next_row + 1, column=1, value=f"WR COUNT: {count_wr}")

            next_row += 2
            sp_df = filtered_df[filtered_df["DUTY TYPE"] == "SP"][["DUTY TYPE", "SIGNON STTN", "SIGNOFF STTN"]]
            ws.cell(row=next_row + 1, column=1, value="DUTY_TYPE")
            ws.cell(row=next_row + 1, column=2, value="SIGN_ON")
            ws.cell(row=next_row + 1, column=3, value="SIGN_OFF")

            oi = next_row + 2
            for row in sp_df.values:
                duty_type, signon, signoff = row
                ws.cell(row=oi, column=1, value=f"{duty_type}")
                ws.cell(row=oi, column=2, value=f"{signon}")
                ws.cell(row=oi, column=3, value=f"{signoff}")
                oi += 1

            wb.save(from_path)

            set_column_widths(from_path)


            crew_ids_set = set()
            for i, j in l:
                if i != j:
                    op = df[(df["SIGNON STTN"] == j) & (df["SIGNOFF STTN"] == i)]
                    crew_ids_set.update(op["CREW ID"].dropna().unique())

            filtered_df = df[
                (df["CREW ID"].isin(crew_ids_set)) &
                (df["SIGNON STTN"].isin(g)) &
                (df["SIGNOFF STTN"].isin(g))
            ]

            count_sp = (filtered_df["DUTY TYPE"] == "SP").sum()
            count_wr = (filtered_df["DUTY TYPE"] == "WR").sum()

            to_path = os.path.join(os.path.dirname(x), f"To{zx}.xlsx")
            filtered_df.to_excel(to_path)

            wb = load_workbook(to_path)
            ws = wb.active
            next_row = ws.max_row + 2
            ws.cell(row=next_row, column=1, value=f"SP COUNT: {count_sp}")
            ws.cell(row=next_row + 1, column=1, value=f"WR COUNT: {count_wr}")

            next_row += 2
            sp_df = filtered_df[filtered_df["DUTY TYPE"] == "SP"][["DUTY TYPE", "SIGNON STTN", "SIGNOFF STTN"]]
            ws.cell(row=next_row + 1, column=1, value="DUTY_TYPE")
            ws.cell(row=next_row + 1, column=2, value="SIGN_ON")
            ws.cell(row=next_row + 1, column=3, value="SIGN_OFF")

            oi = next_row + 2
            for row in sp_df.values:
                duty_type, signon, signoff = row
                ws.cell(row=oi, column=1, value=f"{duty_type}")
                ws.cell(row=oi, column=2, value=f"{signon}")
                ws.cell(row=oi, column=3, value=f"{signoff}")
                oi += 1

            wb.save(from_path)

            set_column_widths(to_path)

        messagebox.showinfo("Success", "All files generated successfully!")

    except Exception as e:
        messagebox.showerror("Processing Error", str(e))


# --- GUI ---
load_stations()

root = tk.Tk()
root.title("Crew Duty Filter")
root.geometry("800x300")

file_path_var = tk.StringVar(value="No file selected")

tk.Button(root, text="Select Excel File", command=select_file).pack(pady=10)
tk.Label(root, textvariable=file_path_var, wraplength=650).pack()
tk.Button(root, text="Generate Filtered Files", command=process_file).pack(pady=20)
tk.Button(root, text="Add Stations",command=Add_station).pack(pady=20)
tk.Button(root, text="View Station",command=View_station).pack(pady=20)

root.mainloop()
